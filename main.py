from sqlalchemy import create_engine
from sqlalchemy import Column
from sqlalchemy import Integer, Text
from sqlalchemy import exists
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
import csv
import os
import json


class Unit:
    """Создание объекта класса Unit (Установка) с сопутствующими
     атрибутами.

     Атрибуты класса:
     all_units -- словарь всех объектов класса Unit. Ключ - имя
        объекта, значение - ссылка на объект.
    """

    all_units = dict()

    def __init__(self, name=None):
        """
        Инициализация класса.

        Атрибуты:
        name -- строка представляющая имя объекта
        __load_max -- int представляющий максимальное значение
            загрузки установки.
        streams_in -- словарь объектов класса Stream (поток). Ключ - имя
            объекта, значение - ссылка на объект.
            Представляет входные потоки.
        streams_out -- словарь объектов класса Stream (поток). Ключ - имя
            объекта, значение - ссылка на объект.
            Представляет выходные потоки.
        """

        self.name = name
        self.__load_max = 0  # атрибут инкапсулирован согласно заданию.
        self.streams_in = dict()
        self.streams_out = dict()
        Unit.all_units[self.name] = self

    def set_load_max(self, value=0):
        self.__load_max = value

    def __repr__(self):
        return f"<{self.__class__.__name__}(name={self.name}," \
               f" __load_max={self.__load_max}," \
               f" streams_in={self.streams_in}," \
               f" streams_out={self.streams_out})>"


class AVTUnit(Unit):
    """Создание объекта класса AVTUnit.

    Наследует атрибуты и методы класса Unit.
    """

    pass


class RerunningUnit(Unit):
    """Создание объекта класса RerunningUnit.

    Наследует атрибуты и методы класса Unit.
    """

    pass


class Stream:
    """Создание объекта класса Stream (Поток) с сопутствующими
     атрибутами.

     Атрибуты класса:
     all_streams -- словарь всех объектов класса Stream. Ключ - имя
        объекта, значение - ссылка на объект.
    """

    all_streams = dict()

    def __init__(self, name=None):
        """
        Инициализация класса.

        Атрибуты:
        name -- строка представляющая имя объекта
        dep_units -- список представляющий имена объектов класса Unit
            для которых данный поток является выходным.
        dst_units -- список представляющий имена объектов класса Unit
            для которых данный поток является входным.
        """

        self.name = name
        self.dep_units = []
        self.dst_units = []
        Stream.all_streams[self.name] = self

    def __repr__(self):
        return f"<Stream(name={self.name}," \
               f" dep_units={self.dep_units}," \
               f" dst_units={self.dst_units})>"


__Base = declarative_base()


class __UnitTable(__Base):
    """Создание объекта класса __UnitTable  с сопутствующими атрибутами.

    Представляет таблицу Unit производственной БД.
    Атрибуты:
    id -- int представляющий идентификатор установки.
    name -- строка представляющая имя установки.
    type -- int представляющий тип установки (0 - АВТ, 1 - Вторичного
        производства).
    """

    __tablename__ = 'unit'
    id = Column(Integer, primary_key=True, nullable=False)
    name = Column(Text)
    type = Column(Integer)

    def __repr__(self):
        return f"<Unit(id={self.id}, name={self.name}, type={self.type})>"


class __StreamTable(__Base):
    """Создание объекта класса __StreamTable  с сопутствующими атрибутами.

    Представляет таблицу Stream производственной БД.
    Атрибуты:
    id -- int представляющий идентификатор потока.
    name -- строка представляющая имя потока.
    """

    __tablename__ = 'stream'
    id = Column(Integer, primary_key=True, nullable=False)
    name = Column(Text)

    def __repr__(self):
        return f"<Stream(id={self.id}, name={self.name})>"


class __UnitMaterialTable(__Base):
    """Создание объекта класса __UnitMaterialTable с сопутствующими
        атрибутами.

    Представляет таблицу Unit_material производственной БД.
    Атрибуты:
    unit_id -- int представляющий идентификатор установки.
    stream_id -- int представляющий идентификатор потока.
    feed_flag -- int определяющий является ли соответствующий поток
        выходным (0) или входным (1) для соответствующей установки
    """

    __tablename__ = 'unit_material'
    unit_id = Column(Integer, primary_key=True)
    stream_id = Column(Integer, primary_key=True)
    feed_flag = Column(Integer)

    def __repr__(self):
        return f"<UnitMaterial(unit_id={self.unit_id}," \
               f" stream_id={self.stream_id}, feed_flag={self.feed_flag})>"


class __LoadMaxTable(__Base):
    """Создание объекта класса __UnitMaterialTable с сопутствующими
        атрибутами.

    Атрибуты:
    unit_id -- int представляющий идентификатор установки.
    value -- int представляющий максимальное значение
        загрузки установки.
    """

    __tablename__ = 'load_max'
    unit_id = Column(Integer, primary_key=True)
    value = Column(Integer)

    def __repr__(self):
        return f"<LoadMax(unit_id={self.unit_id}, value={self.value})>"


def get_units(session, verbose=False):
    """Создает объекты классов AVTUnit и RerunningUnit на основе данных
        производственной БД.

    Аргументы:
    session -- сессия подключения к производственной БД.
    verbose -- флаг вывода результатов в консоль.
    """

    new_unit = None
    query = session.query(__UnitTable.name, __UnitTable.type)
    for instance in query:
        if instance.type == 0:
            new_unit = AVTUnit(instance.name)
        elif instance.type == 1:
            new_unit = RerunningUnit(instance.name)
        if verbose:
            print(new_unit)


def set_load_max(session, verbose=False):
    """Обновляет атрибут __load_max существующих объектов класса Unit,
        AVTUnit, RerunningUnit на основе данных производственной БД.

    Аргументы:
    session -- сессия подключения к производственной БД.
    verbose -- флаг вывода результатов в консоль.
    """

    query = session.query(__UnitTable.name, __LoadMaxTable.value)
    query = query.join(__LoadMaxTable,
                       __UnitTable.id == __LoadMaxTable.unit_id)
    for unit in Unit.all_units.values():
        load_max_value = query.filter(__UnitTable.name == unit.name).\
                                      first().value
        unit.set_load_max(load_max_value)
        if verbose:
            print(unit)


def set_related_streams(verbose=False):
    """Обновляет атрибуты streams_in и streams_out существующих объектов
     класса Unit, AVTUnit, RerunningUnit.

    Аргументы:
    verbose -- флаг вывода результатов в консоль.
    """

    for stream in Stream.all_streams.values():
        for unit in stream.dep_units:
            Unit.all_units[unit].streams_out[stream.name] = stream
        for unit in stream.dst_units:
            Unit.all_units[unit].streams_in[stream.name] = stream
    if verbose:
        for unit in Unit.all_units.values():
            print(unit)


def get_streams(session, verbose=False):
    """Создает объекты класса Stream на основе данных производственной БД.

    Аргументы:
    session -- сессия подключения к производственной БД.
    verbose -- флаг вывода результатов в консоль.
    """

    query = session.query(__StreamTable.name)
    for instance in query:
        new_stream = Stream(instance.name)
        if verbose:
            print(new_stream)


def set_related_units(session, verbose=False):
    """Обновляет атрибуты dst_units и dep_units существующих объектов
     класса Stream на основе данных производственной БД.

    Аргументы:
    session -- сессия подключения к производственной БД.
    verbose -- флаг вывода результатов в консоль.
    """

    for stream in Stream.all_streams.values():
        query = session.query(__StreamTable.name.label('stream_name'),
                              __UnitTable.name.label('unit_name'),
                              __UnitMaterialTable.feed_flag)
        query = query.join(__UnitMaterialTable,
                           __UnitMaterialTable.stream_id == __StreamTable.id)
        query = query.join(__UnitTable,
                           __UnitTable.id == __UnitMaterialTable.unit_id)
        query = query.filter(__StreamTable.name == stream.name)
        query = query.order_by('unit_name')
        for instance in query:
            if instance.feed_flag:
                stream.dst_units.append(instance.unit_name)
            else:
                stream.dep_units.append(instance.unit_name)
        if verbose:
            print(stream)


def task_3(session):
    """
    Решение задачи №3:
        Создать запрос к БД, возвращающий входные потоки установок
        в форме: имя_установки, имя_потока_на_входе.

    """

    query = session.query(__UnitTable.name.label('unit_name'),
                          __StreamTable.name.label('stream_name'))
    query = query.join(__UnitMaterialTable,
                       __UnitMaterialTable.unit_id == __UnitTable.id)
    query = query.join(__StreamTable,
                       __StreamTable.id == __UnitMaterialTable.stream_id)
    query = query.filter(__UnitMaterialTable.feed_flag == 1)
    query = query.order_by('unit_name')
    print()
    for instance in query:
        print(instance.unit_name, instance.stream_name)
    print('Task 3 - DONE')


def task_4(session):
    """
    Решение задачи №4:
        Вывести в .csv потоки, которые не поступают и не выходят ни с
        одной установки в формате: id, имя_потока.

    """

    def write_to_csv(data, path=os.getcwd(), file='task_4.csv'):
        with open(''.join([path, '\\', file]), "w", newline='') as csv_file:
            writer = csv.writer(csv_file, delimiter=';')
            for line in data:
                writer.writerow(line)

    query = session.query(__StreamTable.id, __StreamTable.name)
    sub_query = ~exists().where(__UnitMaterialTable.stream_id ==
                                __StreamTable.id)
    query = query.filter(sub_query)
    file_name = 'task_4.csv'
    write_to_csv(query.all(), file=file_name)
    print(f'\nTask 4 - DONE, check {file_name}')


def task_5():
    """
    Решение задачи №5:
        Вывести в .json потоки, которые поступают на несколько установок
        в формате словаря (ключ - название потока, значение - список установок).

    """

    def write_to_json(_data, path=os.getcwd(), file='task_5.json'):
        with open(''.join([path, '\\', file]), "w") as json_file:
            json.dump(_data, json_file)

    data = dict()
    for stream in Stream.all_streams.values():
        if len(stream.dst_units) > 1:
            data[stream.name] = stream.dst_units
    file_name = 'task_5.json'
    write_to_json(data, file=file_name)
    print(f'\nTask 5 - DONE, check {file_name}')


def task_6():
    """
    Решение задачи №6:
        Создать excel, в котором каждый лист - название установки.
        На листе в первой колонке указать входные потоки, во второй
        выходные (библиотека openpyxl).

    """

    def write_to_xlsx(path=os.getcwd(), file='task_6.xlsx'):
        wb = Workbook()
        for unit in Unit.all_units.values():
            ws = wb.create_sheet(unit.name)
            for y, streams in enumerate([unit.streams_in, unit.streams_out], 1):
                for x, stream in enumerate(streams, 1):
                    ws.cell(row=x, column=y, value=stream)
        wb.remove(wb['Sheet'])
        wb.save(''.join([path, '\\', file]))

    file_name = 'task_6.xlsx'
    write_to_xlsx(file=file_name)
    print(f'\nTask 6 - DONE, check {file_name}')


def main(dialect='sqlite:///', path='db.db'):
    engine = create_engine(''.join([dialect, path]), echo=False)
    Session = sessionmaker(bind=engine)
    with Session() as session:
        get_units(session)
        set_load_max(session)
        get_streams(session)
        set_related_units(session)
        set_related_streams(verbose=True)
        task_3(session)
        task_4(session)
        task_5()
        task_6()


if __name__ == '__main__':
    main()
